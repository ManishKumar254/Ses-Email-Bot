import smtplib
import email.utils
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import time


# Replace sender@example.com with your "From" address.
SENDER = ''
SENDERNAME = ''

# Replace recipient@example.com with a "To" address
RECIPIENT  = ''

# Replace smtp_username with your Amazon SES SMTP user name.
USERNAME_SMTP = ""

# Replace smtp_password with your Amazon SES SMTP password.
PASSWORD_SMTP = ""



HOST = "email-smtp.us-east-2.amazonaws.com"
PORT = 587


SUBJECT = 'Amazon SES Test (Python smtplib)'

# The email body for recipients
BODY_TEXT = ("Amazon SES Test\r\n"
             "This email was sent through the Amazon SES SMTP "
             "Interface using the Python smtplib package."
            )

BODY_HTML = """<html>
<head></head>
<body>
  <h1>Amazon SES SMTP Email Test</h1>
  <p>This email was sent with Amazon SES using the</p>
</body>
</html>
            """

msg = MIMEMultipart('alternative')
msg['Subject'] = SUBJECT
msg['From'] = email.utils.formataddr((SENDERNAME, SENDER))
msg['To'] = RECIPIENT


part1 = MIMEText(BODY_TEXT, 'plain')
part2 = MIMEText(BODY_HTML, 'html')

msg.attach(part1)
msg.attach(part2)


loc = "ses_recipient.xlsx"
wb = openpyxl.load_workbook(loc)
sheet_obj = wb.active
url_list = []
j = 1
while j <= sheet_obj.max_row:
    cell_obj = sheet_obj.cell(row=j, column=1)
    url_list.append(cell_obj.value)
    j = j + 1

print(url_list)
# Try to send the message.
try:
    server = smtplib.SMTP(HOST, PORT)
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.login(USERNAME_SMTP, PASSWORD_SMTP)
    flag=0
    while flag<len(url_list):
        RECIPIENT=url_list[flag]
        if flag%14==0:
            time.sleep(1)
        server.sendmail(SENDER, RECIPIENT, msg.as_string())
        flag=flag+1
    server.close()
# Display an error message if something goes wrong.
except Exception as e:
    print ("Error: ", e)
else:
    print ("Email sent!")
